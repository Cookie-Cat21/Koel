"""CBSL commercial-bank TT exchange rates → macro_series rows.

Source spreadsheet (public statistics page):
https://www.cbsl.gov.lk/en/rates-and-indicators/exchange-rates
File: IF_Buying_Selling_Exchange_Rates.xlsx

Flag: CBSL_FX_ENABLED (default off). Confirm redistribution/attribution
in docs/THIRD_PARTY_DATA.md before enabling in prod.
"""

from __future__ import annotations

import hashlib
import io
import logging
from datetime import UTC, date, datetime
from typing import Any

import httpx

log = logging.getLogger(__name__)

CBSL_FX_XLSX_URL = (
    "https://www.cbsl.gov.lk/sites/default/files/cbslweb_documents/"
    "statistics/sheets/IF_Buying_Selling_Exchange_Rates.xlsx"
)
ATTRIBUTION = "CBSL buying & selling exchange rates (commercial banks TT)"

# series_id → (buy_col_0based_in_row_tuple_with_leading_None, sell_col)
# Row shape from openpyxl:
# (None, date, USD_buy, USD_sell, GBP_buy, GBP_sell, EUR_buy, EUR_sell, ...)
_SERIES_COLS: dict[str, tuple[int, int]] = {
    "USD_LKR": (2, 3),
    "GBP_LKR": (4, 5),
    "EUR_LKR": (6, 7),
}


def _mid(buy: Any, sell: Any) -> float | None:
    try:
        b = float(buy)
        s = float(sell)
    except (TypeError, ValueError):
        return None
    if not (b > 0 and s > 0):
        return None
    return (b + s) / 2.0


def parse_cbsl_fx_xlsx(data: bytes, *, max_rows: int = 500) -> list[dict[str, Any]]:
    """Parse latest ``max_rows`` daily mid rates into upsert dicts."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl required for CBSL FX ingest") from exc

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: list[dict[str, Any]] = []
    raw_hash = hashlib.sha256(data).hexdigest()[:16]

    daily: list[tuple[date, tuple[Any, ...]]] = []
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 4:
            continue
        d = row[1]
        if isinstance(d, datetime):
            d = d.date()
        if not isinstance(d, date):
            continue
        daily.append((d, row))

    daily.sort(key=lambda x: x[0])
    if max_rows > 0:
        daily = daily[-max_rows:]

    for d, row in daily:
        ts = datetime(d.year, d.month, d.day, 12, 0, tzinfo=UTC)
        for series_id, (bi, si) in _SERIES_COLS.items():
            if si >= len(row):
                continue
            mid = _mid(row[bi], row[si])
            if mid is None:
                continue
            out.append(
                {
                    "source": "cbsl_fx",
                    "series_id": series_id,
                    "ts": ts,
                    "value": mid,
                    "unit": "LKR",
                    "as_of_date": d,
                    "attribution": ATTRIBUTION,
                    "raw_hash": raw_hash,
                }
            )
    return out


async def fetch_cbsl_fx_rows(
    *,
    client: httpx.AsyncClient | None = None,
    url: str = CBSL_FX_XLSX_URL,
    max_rows: int = 120,
) -> list[dict[str, Any]]:
    own = client is None
    http = client or httpx.AsyncClient(timeout=60.0, follow_redirects=True)
    try:
        resp = await http.get(
            url,
            headers={"User-Agent": "koel-macro/0.1 (+https://github.com/ArdenoStudio/Koel)"},
        )
        resp.raise_for_status()
        rows = parse_cbsl_fx_xlsx(resp.content, max_rows=max_rows)
        log.info("cbsl_fx: parsed %s points from %s bytes", len(rows), len(resp.content))
        return rows
    finally:
        if own:
            await http.aclose()
