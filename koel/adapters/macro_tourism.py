"""CBSL tourism earnings → ``macro_series`` (Context Hotels / Travel bridge).

Source spreadsheet (External Sector tables):
https://www.cbsl.gov.lk/en/statistics/statistical-tables/External-Sector
File: Earnings from Tourism (2009 January to Latest)

SLTDA publishes country-level arrivals Excel, but their site Terms of Use
restrict redistribution to personal/non-commercial use without written
consent. CBSL republishes monthly tourism *earnings* (USD mn) derived from
SLTDA survey inputs as official public statistics — same intake pattern as
CBSL FX.

Flag: ``SLTDA_TOURISM_ENABLED`` (default off). Series id stays
``TOURISM_ARRIVALS`` for Context wiring; values are earnings (USD mn).
"""

from __future__ import annotations

import hashlib
import io
import logging
import re
from calendar import month_abbr
from datetime import UTC, date, datetime
from typing import Any
from urllib.parse import urljoin

import httpx

log = logging.getLogger(__name__)

ATTRIBUTION = "CBSL earnings from tourism (SLTDA survey inputs)"
SERIES_ID = "TOURISM_ARRIVALS"
UNIT = "USD_mn"

CBSL_EXTERNAL_SECTOR_URL = (
    "https://www.cbsl.gov.lk/en/statistics/statistical-tables/External-Sector"
)
# Dated filename changes when CBSL refreshes the sheet — discover from page.
_FALLBACK_XLSX_URL = (
    "https://www.cbsl.gov.lk/sites/default/files/cbslweb_documents/"
    "statistics/table2.14.1_20260626_e.xlsx"
)

_MONTH_MAP = {
    name.lower(): i
    for i, name in enumerate(month_abbr)
    if name
} | {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


def _parse_year_header(raw: object) -> int | None:
    if isinstance(raw, int) and 1990 <= raw <= 2100:
        return raw
    if isinstance(raw, float) and 1990 <= raw <= 2100:
        return int(raw)
    if isinstance(raw, str):
        m = re.search(r"(19|20)\d{2}", raw)
        if m:
            return int(m.group(0))
    return None


def parse_cbsl_tourism_earnings_xlsx(
    data: bytes,
    *,
    max_points: int = 120,
) -> list[dict[str, Any]]:
    """Parse monthly tourism earnings (USD mn) into upsert dicts."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl required for CBSL tourism ingest") from exc

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw_hash = hashlib.sha256(data).hexdigest()[:16]

    year_row: list[object] | None = None
    year_row_idx: int | None = None
    for r in range(1, min(ws.max_row, 12) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        years = [_parse_year_header(v) for v in vals]
        if sum(1 for y in years if y is not None) >= 2:
            year_row = vals
            year_row_idx = r
            break
    if year_row is None or year_row_idx is None:
        return []

    col_years: dict[int, int] = {}
    for c, cell in enumerate(year_row, start=1):
        y = _parse_year_header(cell)
        if y is not None:
            col_years[c] = y

    points: list[tuple[date, float]] = []
    for r in range(year_row_idx + 1, ws.max_row + 1):
        label = ws.cell(r, 2).value
        if not isinstance(label, str):
            continue
        month = _MONTH_MAP.get(label.strip().lower())
        if month is None:
            continue
        for c, year in col_years.items():
            raw = ws.cell(r, c).value
            if raw is None or raw == "":
                continue
            try:
                value = float(raw)
            except (TypeError, ValueError):
                continue
            if not (value > 0):
                continue
            points.append((date(year, month, 1), value))

    points.sort(key=lambda x: x[0])
    if max_points > 0:
        points = points[-max_points:]

    out: list[dict[str, Any]] = []
    for d, value in points:
        ts = datetime(d.year, d.month, d.day, 12, 0, tzinfo=UTC)
        out.append(
            {
                "source": "cbsl_tourism",
                "series_id": SERIES_ID,
                "ts": ts,
                "value": value,
                "unit": UNIT,
                "as_of_date": d,
                "attribution": ATTRIBUTION,
                "raw_hash": raw_hash,
            }
        )
    return out


def discover_tourism_xlsx_url(html: str, *, base: str = CBSL_EXTERNAL_SECTOR_URL) -> str | None:
    """Pick the Earnings-from-Tourism spreadsheet href from the External Sector page."""
    for m in re.finditer(
        r'<a[^>]+href="([^"]+)"[^>]*>(.*?)</a>',
        html,
        flags=re.I | re.S,
    ):
        href, inner = m.group(1), m.group(2)
        text = re.sub(r"<[^>]+>", "", inner)
        text = re.sub(r"\s+", " ", text).strip()
        if not re.search(r"earnings\s+from\s+tourism", text, re.I):
            continue
        if not re.search(r"\.xlsx?$", href, re.I):
            continue
        return str(urljoin(base, href))
    return None


async def fetch_tourism_rows(
    *,
    client: httpx.AsyncClient | None = None,
    page_url: str = CBSL_EXTERNAL_SECTOR_URL,
    xlsx_url: str | None = None,
    max_points: int = 120,
) -> list[dict[str, Any]]:
    own = client is None
    http = client or httpx.AsyncClient(timeout=60.0, follow_redirects=True)
    headers = {"User-Agent": "koel-macro/0.1 (+https://github.com/ArdenoStudio/Koel)"}
    try:
        url = xlsx_url
        if not url:
            page = await http.get(page_url, headers=headers)
            page.raise_for_status()
            url = discover_tourism_xlsx_url(page.text) or _FALLBACK_XLSX_URL
        resp = await http.get(url, headers=headers)
        resp.raise_for_status()
        rows = parse_cbsl_tourism_earnings_xlsx(resp.content, max_points=max_points)
        log.info(
            "cbsl_tourism: parsed %s points from %s bytes (%s)",
            len(rows),
            len(resp.content),
            url.rsplit("/", 1)[-1],
        )
        return rows
    finally:
        if own:
            await http.aclose()
