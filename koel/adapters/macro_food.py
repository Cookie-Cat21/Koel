"""CBSL CCPI → ``macro_series`` food-pressure Context card.

Source spreadsheet (Real Sector → Prices, Wages and Employment):
https://www.cbsl.gov.lk/en/statistics/statistical-tables/real-sector/prices-wages-employment
File: CCPI and CCPI Core

DCS weekly retail dashboard / microdata carry redistribution restrictions
("All Rights Reserved" / written-agreement clauses). CBSL publishes the
headline Colombo Consumer Price Index (CCPI, base 2021=100) as official
public statistics citing DCS — same intake pattern as CBSL FX.

``FOOD_PRESSURE`` stores the headline CCPI index level (not a grocery SKU
farm). Flag: ``DCS_FOOD_ENABLED`` (default off).
"""

from __future__ import annotations

import hashlib
import io
import logging
import re
from calendar import month_abbr
from datetime import UTC, date, datetime
from typing import Any
from urllib.parse import urljoin

import httpx

log = logging.getLogger(__name__)

ATTRIBUTION = "CBSL CCPI (source: Department of Census and Statistics)"
SERIES_ID = "FOOD_PRESSURE"
UNIT = "index"

CBSL_PRICES_PAGE_URL = (
    "https://www.cbsl.gov.lk/en/statistics/statistical-tables/"
    "real-sector/prices-wages-employment"
)
_FALLBACK_XLSX_URL = (
    "https://www.cbsl.gov.lk/sites/default/files/cbslweb_documents/"
    "statistics/sheets/CCPI_and_CCPI_CORE_20260630_e.xlsx"
)

_MONTH_MAP = {
    name.lower(): i
    for i, name in enumerate(month_abbr)
    if name
} | {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


def _parse_period(label: str) -> date | None:
    """Parse ``2026 June`` / ``June 2026`` → first of month."""
    s = re.sub(r"\s+", " ", label.strip())
    m = re.match(
        r"^(?P<y>(?:19|20)\d{2})\s+(?P<mon>[A-Za-z]+)$",
        s,
    ) or re.match(
        r"^(?P<mon>[A-Za-z]+)\s+(?P<y>(?:19|20)\d{2})$",
        s,
    )
    if not m:
        return None
    month = _MONTH_MAP.get(m.group("mon").lower())
    if month is None:
        return None
    return date(int(m.group("y")), month, 1)


def parse_cbsl_ccpi_xlsx(
    data: bytes,
    *,
    max_points: int = 60,
) -> list[dict[str, Any]]:
    """Parse headline CCPI index column into ``FOOD_PRESSURE`` upserts."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl required for CBSL CCPI ingest") from exc

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw_hash = hashlib.sha256(data).hexdigest()[:16]

    points: list[tuple[date, float]] = []
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 2).value
        idx = ws.cell(r, 3).value
        if not isinstance(label, str):
            continue
        d = _parse_period(label)
        if d is None:
            continue
        if not isinstance(idx, (int, float)):
            continue
        value = float(idx)
        if value <= 0:
            continue
        points.append((d, value))

    points.sort(key=lambda x: x[0])
    if max_points > 0:
        points = points[-max_points:]

    out: list[dict[str, Any]] = []
    for d, value in points:
        ts = datetime(d.year, d.month, d.day, 12, 0, tzinfo=UTC)
        out.append(
            {
                "source": "cbsl_ccpi",
                "series_id": SERIES_ID,
                "ts": ts,
                "value": value,
                "unit": UNIT,
                "as_of_date": d,
                "attribution": ATTRIBUTION,
                "raw_hash": raw_hash,
            }
        )
    return out


def discover_ccpi_xlsx_url(html: str, *, base: str = CBSL_PRICES_PAGE_URL) -> str | None:
    """Pick the CCPI spreadsheet href from the prices/wages page."""
    for m in re.finditer(
        r'<a[^>]+href="([^"]+)"[^>]*>(.*?)</a>',
        html,
        flags=re.I | re.S,
    ):
        href, inner = m.group(1), m.group(2)
        text = re.sub(r"<[^>]+>", "", inner)
        text = re.sub(r"\s+", " ", text).strip()
        if not re.search(r"CCPI\s+and\s+CCPI\s+Core", text, re.I):
            continue
        if not re.search(r"\.xlsx?$", href, re.I):
            continue
        return str(urljoin(base, href))
    return None


async def fetch_food_pressure_rows(
    *,
    client: httpx.AsyncClient | None = None,
    page_url: str = CBSL_PRICES_PAGE_URL,
    xlsx_url: str | None = None,
    max_points: int = 60,
) -> list[dict[str, Any]]:
    own = client is None
    http = client or httpx.AsyncClient(timeout=60.0, follow_redirects=True)
    headers = {"User-Agent": "koel-macro/0.1 (+https://github.com/ArdenoStudio/Koel)"}
    try:
        url = xlsx_url
        if not url:
            page = await http.get(page_url, headers=headers)
            page.raise_for_status()
            url = discover_ccpi_xlsx_url(page.text) or _FALLBACK_XLSX_URL
        resp = await http.get(url, headers=headers)
        resp.raise_for_status()
        rows = parse_cbsl_ccpi_xlsx(resp.content, max_points=max_points)
        log.info(
            "cbsl_ccpi: parsed %s points from %s bytes (%s)",
            len(rows),
            len(resp.content),
            url.rsplit("/", 1)[-1],
        )
        return rows
    finally:
        if own:
            await http.aclose()
